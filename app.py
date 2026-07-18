"""
Psikotes Score Automation — API (v5)
- Improved matching: initials, acronyms, bidirectional word match
- Date normalization: DD/MM/YYYY
- Uniform fields: TGL PEMERIKSAAN and PENDIDIKAN
- Yellow detail list per gugus
- No HTML UI — this is now an internal API consumed only by the Ordinat
  Dashboard (Next.js), which proxies every request with a shared bearer
  token. Never expose this service directly to the public internet.
"""
import os, io, json, uuid, tempfile, threading, re, hmac
from collections import Counter
from functools import wraps
from datetime import datetime
from flask import Flask, request, send_file, jsonify
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
UPLOAD_FOLDER = tempfile.mkdtemp()
jobs = {}
# Each finished job retains a loaded workbook + its output file. Without a cap
# this grows unbounded on a long-running server, so we evict the oldest
# terminal jobs past this limit (see _evict_old_jobs).
MAX_JOBS = 30


def require_service_token(fn):
    """Gate a route behind Authorization: Bearer <RECAP_SERVICE_TOKEN>.
    Fails closed — an unset token never authorizes anything. This is the same
    shared secret the dashboard uses for its Flask -> Next.js webhook, reused
    here for the Next.js -> Flask direction."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        token = os.environ.get('RECAP_SERVICE_TOKEN')
        auth = request.headers.get('Authorization', '')
        if not token or not auth.startswith('Bearer ') or not hmac.compare_digest(auth[7:], token):
            return jsonify({'error': 'Unauthorized'}), 401
        return fn(*args, **kwargs)
    return wrapper

# Skor kecocokan nama di bawah ini butuh konfirmasi manual sebelum ditulis ke Rekap.
AUTO_CONFIDENCE = 0.90

# Floor below which a leftover student's "closest roster name" is just noise
# (coincidental partial overlap, not a plausible same-person typo) — e.g. a
# junk RAW row scoring 15-50% against an unrelated name. Below this, the
# closest_candidate is still shown for transparency but not offered as a
# one-click-confirmable match in the review UI.
LEFTOVER_CANDIDATE_MIN = 0.55


# ── Name Matching ──

def normalize_name(name):
    if pd.isna(name): return ""
    n = str(name).strip().upper()
    for ch in ["\u2018","\u2019","'",'"']: n = n.replace(ch,"'")
    n = n.replace("."," ").replace("-"," ")
    while "  " in n: n = n.replace("  "," ")
    return n.strip()

def _ordered_word_match(short_w, long_w):
    """Match words in order. A short token (<=4 chars) that's an exact prefix
    of a long-side word counts as a partial (initial/abbreviation) match —
    covers both single-letter initials ("M" -> MUHAMMAD) and the common
    Indonesian 3-4 letter prefix abbreviations ("Moh"/"Muh"/"Moch" -> MUHAMMAD).
    A token that exactly equals two CONSECUTIVE long-side words concatenated
    also counts as a full match — covers a compact-vs-split spelling of the
    same name ("AZZAHRA" == "AZ"+"ZAHRA"), which the prefix rule above can't
    reach since it only ever compares one short token to one long word."""
    li, matched = 0, 0
    for sw in short_w:
        for j in range(li, len(long_w)):
            lw = long_w[j]
            if sw == lw: matched += 1; li = j+1; break
            elif len(sw) <= 4 and lw.startswith(sw): matched += 0.8; li = j+1; break
            elif j + 1 < len(long_w) and sw == lw + long_w[j+1]: matched += 1; li = j+2; break
    return matched

def calc_match(a, b):
    if not a or not b: return 0
    if a == b: return 1.0
    s = SequenceMatcher(None, a, b).ratio()
    # Substring check (min 4 chars to avoid short false positives)
    if len(a)>=4 and a in b: s = max(s, len(a)/len(b), 0.82)
    if len(b)>=4 and b in a: s = max(s, len(b)/len(a), 0.82)
    wa, wb = a.split(), b.split()
    if len(wa) >= 2 and len(wb) >= 2:
        # 1) Ordered word match with initials/prefix abbreviations
        #    (M->MUHAMMAD, MOH->MUHAMMAD, MUH->MUHAMMAD, P->PUTRA)
        for sw, lw in [(wa,wb),(wb,wa)]:
            m = _ordered_word_match(sw, lw)
            if len(sw) > 0:
                ratio = m / len(sw)
                # Scaled to word count, not a fixed m>=2 — a fixed floor was
                # unreachable for 2-word names where one word is a single
                # initial (e.g. "M. IRFAN" vs "MUKHAMAD IRFAN": max possible
                # m = 0.8 (initial) + 1.0 (exact) = 1.8, never >= 2), so this
                # very common Indonesian abbreviation pattern (Muhammad/
                # Mohammad/Muhamad -> "M.") was silently unmatchable no
                # matter how confident the rest of the name was.
                if ratio >= 0.6 and m >= len(sw) - 0.5: s = max(s, ratio * 0.92)
        # 2) Acronym detection — strict rule:
        #    A token (2-6 chars) is an acronym ONLY if each character is the
        #    FIRST LETTER of a CONSECUTIVE word on the other side.
        #    e.g. "SPS" = S(haista) P(utri) S(etiawan)
        #    Remaining non-acronym words must also match.
        for all_w, check_w in [(wa,wb),(wb,wa)]:
            for token in all_w:
                if len(token) < 2 or len(token) > 6: continue
                if token in check_w: continue  # already a real word, skip
                for start in range(len(check_w)):
                    needed = len(token)
                    if start + needed > len(check_w): break
                    candidate = check_w[start:start+needed]
                    initials = ''.join(w[0] for w in candidate)
                    if initials == token:
                        # Acronym match! Score based on remaining word matches
                        rest_tokens = [w for w in all_w if w != token]
                        rest_targets = [w for w in check_w if w not in candidate]
                        if not rest_tokens:
                            s = max(s, 0.85)
                        else:
                            rm = sum(1 for rt in rest_tokens if rt in rest_targets)
                            tr = (rm + 1) / (len(rest_tokens) + 1)
                            if tr > 0.5: s = max(s, tr * 0.90)
                        break
    return s


# ── Date Normalization ──

def normalize_date(val):
    """Convert any date value to DD/MM/YYYY string."""
    if pd.isna(val) or val is None: return None
    if isinstance(val, (datetime, pd.Timestamp)):
        y = val.year
        if y < 100: y += 2000
        return f"{val.day:02d}/{val.month:02d}/{y:04d}" if 1900 <= y <= 2100 else None
    s = str(val).strip()
    if not s: return None
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        a, b, ys = int(m.group(1)), int(m.group(2)), m.group(3)
        y = int(ys)
        if y < 1900: y = 2000 + int(ys[-2:])  # "0209"->2009, "0008"->2008
        if a > 12 and b <= 12: a, b = b, a
        if 1<=a<=12 and 1<=b<=31: return f"{b:02d}/{a:02d}/{y:04d}"
        if 1<=b<=12 and 1<=a<=31: return f"{a:02d}/{b:02d}/{y:04d}"
        return s
    m2 = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{2,4})$', s)
    if m2:
        a, b, y = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
        if y < 100: y += 2000
        if a > 12 and b <= 12: a, b = b, a
        return f"{b:02d}/{a:02d}/{y:04d}" if 1<=a<=12 else s
    try:
        dt = pd.to_datetime(val)
        y = dt.year
        if y < 100: y += 2000
        return f"{dt.day:02d}/{dt.month:02d}/{y:04d}" if 1900<=y<=2100 else s
    except: return s


# ── Sheet Detection ──

def fast_detect_sheets(raw_path):
    wb = load_workbook(raw_path, read_only=True, data_only=True)
    subtes, other = [], []
    for sn in wb.sheetnames:
        ws = wb[sn]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        hdrs = [str(h).strip() if h else "" for h in row1]
        if 'Score' in hdrs and 'NAMA LENGKAP' in hdrs and 'KELAS' in hdrs: subtes.append(sn)
        else: other.append(sn)
    wb.close()
    return subtes, other

def read_raw_minimal(raw_path, sheet_name):
    needed = ['Timestamp','Score','NAMA LENGKAP','KELAS',
              'JENIS KELAMIN','TEMPAT LAHIR','TANGGAL LAHIR','TANGGAL TES','PENDIDIKAN ']
    return pd.read_excel(raw_path, sheet_name=sheet_name, usecols=lambda c: c in needed)

def detect_kelas_fmt(vals):
    for v in vals:
        s = str(v)
        if s.startswith('KELAS X'): return 'KELAS X'
        if 'XI.' in s: return 'XI.'
    return 'X'

def extract_kelas_num(kstr, fmt):
    try:
        s = str(kstr).strip()
        if fmt == 'KELAS X': return int(s.replace('KELAS X','').strip())
        if fmt == 'XI.': return int(s.replace('XI.','').strip())
        return int(s.replace('X','').strip())
    except: return 0

def find_id_sheet(raw_path, subtes_sheets):
    wb = load_workbook(raw_path, read_only=True, data_only=True)
    res = None
    for sn in subtes_sheets:
        ws = wb[sn]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        hdrs = [str(h).upper().strip() if h else "" for h in row1]
        if any('JENIS KELAMIN' in h for h in hdrs): res = sn; break
    wb.close()
    return res or ('SE' if 'SE' in subtes_sheets else subtes_sheets[0] if subtes_sheets else None)


# ── Template Layout ──
# The v2 Rekap template uses a two-row header: row 1 holds merged section labels
# (BIODATA / EPPS / RIASEC / GAYA BELAJAR) and row 2 holds the field headers
# (NOMOR, NAMA, JK, the Score-subtes labels, and per-question numbers), with data
# from row 3. Older single-row templates put field headers in row 1 and data from
# row 2. Everything below auto-detects which layout it's looking at so the
# existing Score-copy path keeps working unchanged on both.

# Answer-type sections whose columns are a numbered question block, not a single
# score cell. Keys are the row-1 section labels as they appear in the template.
ANSWER_SECTION_LABELS = {'EPPS', 'RIASEC', 'GAYA BELAJAR'}

def detect_header_layout(ws):
    """Return (header_row, data_start) by locating the row that holds 'NAMA'.
    Two-row template -> (2, 3); legacy single-row template -> (1, 2)."""
    for hr in (1, 2):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=hr, column=col).value
            if v is not None and str(v).strip().upper() == 'NAMA':
                return hr, hr + 1
    return 1, 2  # fallback: legacy behavior

def read_field_headers(ws, header_row):
    """Map field-label -> column from header_row, skipping the numeric
    per-question cells (1,2,3,...) that belong to answer-section blocks."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            continue  # question number, not a field label
        sv = str(v).strip()
        if not sv or sv.isdigit():
            continue
        headers[sv] = col
    return headers

def map_answer_sections(ws, header_row):
    """Return {section_label: {question_num(int): column}} for the answer-type
    blocks. Section spans come from the merged label band in row 1; question
    numbers come from header_row. Empty on legacy templates that lack these
    blocks."""
    label_cols = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None and str(v).strip().upper() in ANSWER_SECTION_LABELS:
            label_cols[col] = str(v).strip().upper()
    if not label_cols:
        return {}
    # resolve each label's column span from the merged range it anchors
    spans = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.min_col in label_cols:
            spans[label_cols[mr.min_col]] = (mr.min_col, mr.max_col)
    # non-merged single-cell labels: span to the column before the next label/end
    starts = sorted(label_cols)
    for i, c0 in enumerate(starts):
        lbl = label_cols[c0]
        if lbl in spans:
            continue
        c1 = (starts[i + 1] - 1) if i + 1 < len(starts) else ws.max_column
        spans[lbl] = (c0, c1)
    sections = {}
    for lbl, (c0, c1) in spans.items():
        qmap = {}
        for col in range(c0, c1 + 1):
            hv = ws.cell(row=header_row, column=col).value
            if isinstance(hv, (int, float)) and not isinstance(hv, bool):
                qmap[int(hv)] = col
            elif hv is not None and str(hv).strip().isdigit():
                qmap[int(str(hv).strip())] = col
        if qmap:
            sections[lbl] = qmap
    return sections


# ── Answer-choice subtests (EPPS / RIASEC / GB) ──
# These RAW sheets carry one answer string per question (no single Score). We
# name-match the student exactly like a normal subtest, then write the parsed
# choice token ("A." / "b." / "Ya") into each question column of the matching
# REKAP section block — no value copy, no score computation.

# RAW sheet name -> REKAP section label (row-1 merged label).
RAW_TO_SECTION = {'EPPS': 'EPPS', 'RIASEC': 'RIASEC', 'GB': 'GAYA BELAJAR'}

def extract_choice(val):
    """Take the leading choice token from an answer, dot included, case
    preserved: "A. Saya suka..." -> "A.", "b. Mendengarkan..." -> "b.".
    Values without a letter+dot prefix (RIASEC's "Ya"/"Tidak") pass through
    trimmed. Blank/NaN -> None (write nothing / empty cell)."""
    if pd.isna(val): return None
    s = str(val).strip()
    if not s: return None
    m = re.match(r'^([A-Za-z]+)\.', s)
    return m.group(0) if m else s

def detect_answer_sheets(raw_path):
    """RAW sheets shaped as per-question answers: NAMA LENGKAP + KELAS, no
    Score, followed by numbered question columns. Returns [(sheet, section)]."""
    wb = load_workbook(raw_path, read_only=True, data_only=True)
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        hdrs = [str(h).strip() if h else "" for h in row1]
        up = [h.upper() for h in hdrs]
        if 'NAMA LENGKAP' not in up or 'KELAS' not in up: continue
        if 'SCORE' in up: continue  # score-type -> existing path
        if not any(re.match(r'^\d+\.', h) for h in hdrs): continue
        section = RAW_TO_SECTION.get(sn.strip().upper(), sn.strip().upper())
        out.append((sn, section))
    wb.close()
    return out

def read_raw_answers(raw_path, sheet_name):
    """Return (df, qcols): NAMA LENGKAP + KELAS + every numbered question
    column in sheet order, plus a normalized-name column."""
    df = pd.read_excel(raw_path, sheet_name=sheet_name)
    qcols = [c for c in df.columns if re.match(r'^\d+\.', str(c).strip())]
    keep = [c for c in ['NAMA LENGKAP', 'KELAS'] + qcols if c in df.columns]
    out = df[keep].copy()
    out['_norm'] = out['NAMA LENGKAP'].apply(normalize_name)
    return out, qcols

def _best_name_match(gn, mt, pool):
    """Highest calc_match over a RAW pool; a confident manual override wins."""
    best_s, best_i = 0, None
    for ri, r in pool.iterrows():
        s = calc_match(gn, r['_norm'])
        if mt and calc_match(mt, r['_norm']) > 0.9: s = 0.95
        if s > best_s: best_s, best_i = s, ri
    return best_s, best_i

# Quality-gate thresholds ("ngawur" detection) — tunable.
STRAIGHTLINE_FRAC = 0.90   # one choice dominating this fraction = suspect
SPARSE_MIN_FRAC = 0.50     # answered below this fraction = suspect

def assess_answer_quality(answers):
    """Flag careless/nonsense ('ngawur') answer sheets. `answers` is the list of
    parsed tokens for one student. Returns (reason, severity, detail) or None if
    the response looks genuine."""
    n_total = len(answers)
    if n_total == 0: return None
    vals = [a for a in answers if a not in (None, '')]
    n_ans = len(vals)
    if n_ans == 0:
        return ('C3_blank', 'warn', 'Lembar jawaban kosong')
    top, cnt = Counter(vals).most_common(1)[0]
    if n_ans >= 5 and cnt / n_ans >= STRAIGHTLINE_FRAC:
        return ('C1_straightlining', 'block',
                f'{int(cnt/n_ans*100)}% jawaban sama ("{top}") — kemungkinan ngawur')
    if n_ans / n_total < SPARSE_MIN_FRAC:
        return ('C2_sparse', 'warn', f'hanya {n_ans}/{n_total} soal terisi')
    return None

GE_ZERO_MIN_ANSWERS = 5     # GE=0 despite this many filled answers = suspect

def assess_ge(score_val, row_series, qcols):
    """Flag a GE score of 0 that came from a student who *did* answer — usually
    means gibberish/typo'd free-text (nothing matched the rubric), worth a look.
    Returns (reason, severity, detail) or None."""
    if score_val != 0: return None
    n_ans = sum(1 for c in qcols
                if not pd.isna(row_series.get(c)) and str(row_series.get(c)).strip())
    if n_ans >= GE_ZERO_MIN_ANSWERS:
        return ('V3_ge_zero', 'block',
                f'GE = 0 padahal {n_ans} soal terisi — kemungkinan ngawur/typo')
    return None


# ── GE computed scoring ──
# GE answers are free text, not a score. Each question has an ordered
# keyword→weight rubric (from RUMUS GE.xlsx, extracted once into
# data/ge_rubric.json, keyed by question number "61".."76"). Per question the
# first exact keyword match wins (best synonym=2, looser term=1, else 0); the GE
# value is the sum (0–32). The RAW "Score" column is ignored — recomputed here.

GE_RUBRIC = {}
try:
    _ge_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'ge_rubric.json')
    with open(_ge_path, encoding='utf-8') as _f:
        GE_RUBRIC = json.load(_f)
except Exception:
    GE_RUBRIC = {}  # GE falls back to writing nothing computable if rubric missing

def _norm_ge_answer(val):
    """Lowercase + trim + collapse whitespace, replicating Excel's `=` compare."""
    if pd.isna(val): return ""
    return re.sub(r'\s+', ' ', str(val).strip().lower())

def read_ge(raw_path, sheet_name):
    """Return (df, qcols): Score + NAMA LENGKAP + KELAS + the numbered GE answer
    columns (dropped by read_raw_minimal), so the score can be recomputed."""
    df = pd.read_excel(raw_path, sheet_name=sheet_name)
    qcols = [c for c in df.columns if re.match(r'^\d+\.', str(c).strip())]
    keep = [c for c in ['Score', 'NAMA LENGKAP', 'KELAS'] + qcols if c in df.columns]
    return df[keep].copy(), qcols

def score_ge(row_series, qcols):
    """Sum GE rubric weights across the answer columns; first exact keyword match
    per question wins, unmatched = 0. Returns int 0–32."""
    total = 0
    for col in qcols:
        m = re.match(r'^(\d+)\.', str(col).strip())
        if not m: continue
        rubric = GE_RUBRIC.get(m.group(1))
        if not rubric: continue
        ans = _norm_ge_answer(row_series.get(col))
        if not ans: continue
        for entry in rubric:  # ordered → first match wins
            if ans == entry['kw']:
                total += entry['w']; break
    return total

def value_for(subtes, row_series, ge_qcols):
    """Value to write for a matched RAW row: computed sum for GE, copied Score
    otherwise. Keeps the existing value-copy behavior for every other subtest."""
    if str(subtes).strip().upper() == 'GE':
        return score_ge(row_series, ge_qcols)
    return row_series.get('Score')


# ── Background Job ──

def _remove_file(path):
    if path and os.path.exists(path):
        try: os.remove(path)
        except OSError: pass

def _free_job_memory(job):
    """Release a terminal job's heavy resources: the loaded openpyxl workbook +
    raw DataFrames (held in _ctx) and the input files on disk. The output file
    and the log survive — they're still needed for /download and /status."""
    job.pop('_ctx', None)  # drops the workbook + all_raw/all_proktor
    _remove_file(job.pop('_raw_path', None))
    _remove_file(job.pop('_rekap_path', None))

def _evict_old_jobs():
    """Cap memory/disk by dropping the oldest terminal (done/error) jobs, and
    their output files, once we exceed MAX_JOBS. Active and awaiting_review
    jobs are never evicted. `jobs` preserves insertion order, so the terminal
    list below is oldest-first."""
    if len(jobs) <= MAX_JOBS: return
    terminal = [jid for jid, j in jobs.items() if j.get('status') in ('done', 'error')]
    for jid in terminal:
        if len(jobs) <= MAX_JOBS: break
        j = jobs.pop(jid, None)
        if j:
            _remove_file(os.path.join(UPLOAD_FOLDER, j['download_filename']) if j.get('download_filename') else None)

def queue_or_write(ws, row_idx, col_idx, sv, best_s, gnum, subtes, gn_display, matched_name, source, log, pending, flag=None):
    """Write score directly if high-confidence and clean; otherwise queue it for
    manual review. `flag` is an optional (reason, severity, detail) anomaly (e.g.
    GE zero-signal) that forces review even on a confident name match.
    Returns True if the row is considered handled (found), either way."""
    if pd.isna(sv): return False
    # Scores come from Google Forms responses — a stray non-numeric cell must
    # skip just this row (it'll surface as missing/yellow), never crash the job.
    try:
        score_val = int(float(sv))
    except (ValueError, TypeError):
        return False
    if flag is None and best_s >= AUTO_CONFIDENCE:
        ws.cell(row=row_idx, column=col_idx, value=score_val)
        log['total_scores'] += 1
        if source != 'own': log['cross_kelas'] += 1
    else:
        reason, severity, detail = flag if flag else (
            'N1_borderline', 'block', f'skor kecocokan nama {round(best_s, 3)}')
        pending.append({
            'id': len(pending), 'gugus': gnum, 'subtes': subtes,
            'nama_rekap': gn_display, 'nama_raw': str(matched_name),
            'score': round(best_s, 3), 'source': source, 'kind': 'value',
            'reason': reason, 'severity': severity, 'detail': detail,
            '_row_idx': row_idx, '_col_idx': col_idx, '_value': score_val,
        })
    return True

def _apply_leftover_confirm(wb, raw_path, item, subtes_list, headers, answer_sections, ge_qcols):
    """Write a confirmed subset of a leftover student's subtests into their
    matched candidate's roster row (item['_candidate_gugus']/['_candidate_row'],
    set only when 'confirmable' — see the leftover-check pass in process_job).

    Re-reads raw_path directly for just the requested sheets rather than
    caching every RAW row up front — leftover confirmation is the rare,
    manually-triggered path, not the per-cell hot loop the rest of this file
    optimizes for. raw_path is still on disk at this point: it's only removed
    by _free_job_memory, which runs after finalize_job, which this function's
    caller (submit_review) always runs strictly before.

    Returns True if at least one subtest was actually written."""
    gnum, row_idx, k = item.get('_candidate_gugus'), item.get('_candidate_row'), item.get('_norm_key')
    if gnum is None or row_idx is None or not k: return False
    ws = wb[f'Gugus {gnum}']
    wrote = False
    for sn in subtes_list:
        sn = str(sn).strip().upper()
        if sn in headers:
            # Score-type subtest (SE/WA/AN/GE/RA/ZR/FA/WU/ME/...): one cell.
            if sn == 'GE':
                df, gqc = read_ge(raw_path, sn)
            else:
                df, gqc = read_raw_minimal(raw_path, sn), []
            df = df.copy()
            df['_norm'] = df['NAMA LENGKAP'].apply(normalize_name)
            match = df[df['_norm'] == k]
            if match.empty: continue
            row = match.iloc[0]
            sv = value_for(sn, row, gqc)
            if pd.isna(sv): continue
            try: score_val = int(float(sv))
            except (ValueError, TypeError): continue
            ws.cell(row=row_idx, column=headers[sn], value=score_val)
            wrote = True
        else:
            # Answer-choice subtest (EPPS/RIASEC/GB): a whole question block.
            section = RAW_TO_SECTION.get(sn, sn)
            qmap = answer_sections.get(section)
            if not qmap: continue
            found_sheet = next((asn for asn, _ in detect_answer_sheets(raw_path)
                                 if asn.strip().upper() == sn), None)
            if not found_sheet: continue
            adf, qcols = read_raw_answers(raw_path, found_sheet)
            match = adf[adf['_norm'] == k]
            if match.empty: continue
            row = match.iloc[0]
            n_write = min(len(qcols), len(qmap))
            answers = [extract_choice(row[qcols[i]]) for i in range(n_write)]
            targets = [qmap.get(i+1) for i in range(n_write)]
            for col, val in zip(targets, answers):
                if col and val is not None:
                    ws.cell(row=row_idx, column=col, value=val)
                    wrote = True
    return wrote

def process_job(job_id, raw_path, rekap_path, overrides, threshold, tgl_pemeriksaan, pendidikan):
    job = jobs[job_id]
    try:
        job['status'] = 'Mendeteksi sheet...'; job['progress'] = 3
        norm_ov = {normalize_name(k): (normalize_name(v) if v else None) for k,v in overrides.items()}
        subtes_sheets, _ = fast_detect_sheets(raw_path)
        if not subtes_sheets:
            job['status'] = 'error'; job['error'] = 'Tidak ada sheet subtes di RAW'; return

        job['progress'] = 8
        all_raw = {}; all_proktor = {}; kelas_fmt = None; ge_qcols = []
        raw_roster = {}  # _norm -> {'nama','kelas','sheets'}: every real RAW student
                         # (+ which sheets they appear in), for the leftover check
        for i, sn in enumerate(subtes_sheets):
            job['status'] = f'Membaca {sn}...'
            if sn.strip().upper() == 'GE':
                raw, ge_qcols = read_ge(raw_path, sn)  # keep answer cols to recompute
            else:
                raw = read_raw_minimal(raw_path, sn)
            raw['_norm'] = raw['NAMA LENGKAP'].apply(normalize_name)
            # Split: normal kelas vs Proktor entries
            is_proktor = raw['KELAS'].astype(str).str.contains('Proktor', na=False)
            proktor_df = raw[is_proktor].drop_duplicates(subset=['_norm'], keep='first')
            normal_df = raw[~is_proktor].drop_duplicates(subset=['_norm','KELAS'], keep='first')
            all_raw[sn] = normal_df
            all_proktor[sn] = proktor_df
            for k, nm, kl in zip(normal_df['_norm'], normal_df['NAMA LENGKAP'], normal_df['KELAS']):
                if not k: continue
                raw_roster.setdefault(k, {'nama': nm, 'kelas': kl, 'sheets': set()})['sheets'].add(sn)
            if kelas_fmt is None:
                kv = [v for v in normal_df['KELAS'].dropna().unique()]
                kelas_fmt = detect_kelas_fmt(kv)
            job['progress'] = 8 + int(22*(i+1)/len(subtes_sheets))

        job['status'] = 'Membaca Rekap...'; job['progress'] = 32
        wb = load_workbook(rekap_path)
        gsheets = sorted([s for s in wb.sheetnames if s.lower().startswith('gugus')],
                          key=lambda x: int(''.join(filter(str.isdigit, x)) or 0))
        if not gsheets:
            job['status'] = 'error'; job['error'] = 'Tidak ada sheet Gugus'; return

        ws0 = wb[gsheets[0]]
        header_row, data_start = detect_header_layout(ws0)
        headers = read_field_headers(ws0, header_row)
        answer_sections = map_answer_sections(ws0, header_row)
        name_col = headers.get('NAMA', 2)
        rekap_subtes = [s for s in subtes_sheets if s in headers]

        gugus_rows, gugus_all_names, gugus_nums = {}, {}, []
        for sn in gsheets:
            try: gnum = int(''.join(filter(str.isdigit, sn)))
            except: continue
            gugus_nums.append(gnum)
            ws = wb[sn]
            gugus_rows[gnum] = {}
            for row in range(data_start, ws.max_row+1):
                cv = ws.cell(row=row, column=name_col).value
                if cv: gugus_rows[gnum][row] = normalize_name(cv)
            gugus_all_names[gnum] = list(gugus_rows[gnum].values())

        job['progress'] = 38; job['status'] = 'Mencocokkan nama...'
        log = {'total_scores':0,'cross_kelas':0,'identity_filled':0,'yellow_count':0,
               'per_gugus':{},'unmatched_summary':[],'yellow_detail':[],
               'subtes_detected':subtes_sheets,'rekap_subtes':rekap_subtes}

        total_steps = len(rekap_subtes) * len(gugus_nums); step = 0
        pending = []
        for subtes in rekap_subtes:
            raw = all_raw[subtes]; proktor = all_proktor[subtes]; col_idx = headers[subtes]
            for gnum in gugus_nums:
                if gnum not in gugus_rows: step+=1; continue
                kstr = f'{kelas_fmt} {gnum}' if kelas_fmt != 'XI.' else f'XI.{gnum}'
                ws = wb[f'Gugus {gnum}']
                raw_own = raw[raw['KELAS']==kstr]
                for row_idx, gn in gugus_rows[gnum].items():
                    found = False; mt = norm_ov.get(gn)
                    # Step 1: own kelas
                    best_s, best_i = 0, None
                    for ri, r in raw_own.iterrows():
                        s = calc_match(gn, r['_norm'])
                        if mt and calc_match(mt, r['_norm'])>0.9: s=0.95
                        if s>best_s: best_s,best_i=s,ri
                    if best_s>=threshold and best_i is not None:
                        mrow = raw.loc[best_i]
                        sv = value_for(subtes, mrow, ge_qcols)
                        flag = assess_ge(sv, mrow, ge_qcols) if str(subtes).strip().upper()=='GE' else None
                        found = queue_or_write(ws, row_idx, col_idx, sv, best_s, gnum, subtes, gn, mrow['NAMA LENGKAP'], 'own', log, pending, flag)
                    # Step 2: cross-kelas
                    if not found:
                        raw_other = raw[raw['KELAS']!=kstr]; best_s,best_i=0,None
                        for ri, r in raw_other.iterrows():
                            s = calc_match(gn, r['_norm'])
                            if mt and calc_match(mt, r['_norm'])>0.9: s=0.95
                            if s>best_s: best_s,best_i=s,ri
                        if best_s>=threshold and best_i is not None:
                            rr = raw.loc[best_i]; rg = extract_kelas_num(rr['KELAS'], kelas_fmt)
                            claimed = rg in gugus_all_names and any(calc_match(og, rr['_norm'])>=threshold for og in gugus_all_names[rg])
                            if not claimed:
                                sv = value_for(subtes, rr, ge_qcols)
                                flag = assess_ge(sv, rr, ge_qcols) if str(subtes).strip().upper()=='GE' else None
                                found = queue_or_write(ws, row_idx, col_idx, sv, best_s, gnum, subtes, gn, rr['NAMA LENGKAP'], 'cross', log, pending, flag)
                    # Step 3: search Proktor entries (students who picked wrong kelas option)
                    if not found and len(proktor) > 0:
                        best_s, best_i = 0, None
                        for ri, r in proktor.iterrows():
                            s = calc_match(gn, r['_norm'])
                            if mt and calc_match(mt, r['_norm'])>0.9: s=0.95
                            if s>best_s: best_s,best_i=s,ri
                        if best_s>=threshold and best_i is not None:
                            mrow = proktor.loc[best_i]
                            sv = value_for(subtes, mrow, ge_qcols)
                            flag = assess_ge(sv, mrow, ge_qcols) if str(subtes).strip().upper()=='GE' else None
                            found = queue_or_write(ws, row_idx, col_idx, sv, best_s, gnum, subtes, gn, mrow['NAMA LENGKAP'], 'proktor', log, pending, flag)
                step += 1
                job['progress'] = min(38 + int(40*step/max(total_steps,1)), 78)
            job['status'] = f'Mencocokkan {subtes}...'

        # ── Answer-choice subtests (EPPS / RIASEC / GB) ──
        # Same 3-tier name match as above, but write parsed choice tokens across
        # the section's question-number columns instead of one Score cell.
        if answer_sections:
            job['status'] = 'Mencocokkan jawaban...'
            for sn, section in detect_answer_sheets(raw_path):
                if section not in answer_sections: continue
                adf, qcols = read_raw_answers(raw_path, sn)
                is_prok = adf['KELAS'].astype(str).str.contains('Proktor', na=False)
                aprok = adf[is_prok].drop_duplicates(subset=['_norm'], keep='first')
                anorm = adf[~is_prok].drop_duplicates(subset=['_norm','KELAS'], keep='first')
                for k, nm, kl in zip(anorm['_norm'], anorm['NAMA LENGKAP'], anorm['KELAS']):
                    if not k: continue
                    raw_roster.setdefault(k, {'nama': nm, 'kelas': kl, 'sheets': set()})['sheets'].add(sn)
                qmap = answer_sections[section]
                n_write = min(len(qcols), len(qmap))
                for gnum in gugus_nums:
                    if gnum not in gugus_rows: continue
                    kstr = f'{kelas_fmt} {gnum}' if kelas_fmt != 'XI.' else f'XI.{gnum}'
                    ws = wb[f'Gugus {gnum}']
                    own = anorm[anorm['KELAS']==kstr]; other = anorm[anorm['KELAS']!=kstr]
                    for row_idx, gn in gugus_rows[gnum].items():
                        mt = norm_ov.get(gn)
                        best_s, best_i, src, pool = 0, None, None, None
                        bs, bi = _best_name_match(gn, mt, own)
                        if bs>=threshold and bi is not None:
                            best_s, best_i, src, pool = bs, bi, 'own', anorm
                        if best_i is None:
                            bs, bi = _best_name_match(gn, mt, other)
                            if bs>=threshold and bi is not None:
                                rr = anorm.loc[bi]; rg = extract_kelas_num(rr['KELAS'], kelas_fmt)
                                claimed = rg in gugus_all_names and any(calc_match(og, rr['_norm'])>=threshold for og in gugus_all_names[rg])
                                if not claimed: best_s, best_i, src, pool = bs, bi, 'cross', anorm
                        if best_i is None and len(aprok)>0:
                            bs, bi = _best_name_match(gn, mt, aprok)
                            if bs>=threshold and bi is not None:
                                best_s, best_i, src, pool = bs, bi, 'proktor', aprok
                        if best_i is None: continue
                        row = pool.loc[best_i]
                        answers = [extract_choice(row[qcols[i]]) for i in range(n_write)]
                        targets = [qmap.get(i+1) for i in range(n_write)]
                        # Quality gate: even a confident name match must be
                        # confirmed if the content looks ngawur.
                        flag = assess_answer_quality(answers)
                        if flag is None and best_s >= AUTO_CONFIDENCE:
                            for col, val in zip(targets, answers):
                                if col and val is not None: ws.cell(row=row_idx, column=col, value=val)
                            log['total_scores'] += 1
                            if src != 'own': log['cross_kelas'] += 1
                        else:
                            reason, severity, detail = flag if flag else (
                                'N1_borderline', 'block', f'skor kecocokan nama {round(best_s,3)}')
                            pending.append({
                                'id': len(pending), 'gugus': gnum, 'subtes': section,
                                'nama_rekap': gn, 'nama_raw': str(row['NAMA LENGKAP']),
                                'score': round(best_s,3), 'source': src, 'kind': 'answer_choice',
                                'reason': reason, 'severity': severity, 'detail': detail,
                                '_row_idx': row_idx, '_answers': answers, '_targets': targets,
                            })

        # ── Leftover check (N7): RAW student missing from the REKAP roster ──
        # A RAW student counts as "missing" only if their class maps to a gugus
        # that exists in this REKAP yet no roster name matches them — otherwise
        # they belong to a class this REKAP simply doesn't cover (not an anomaly).
        gugus_set = set(gugus_nums)
        leftover = []
        for k, info in raw_roster.items():
            display, kelas, sheets = info['nama'], info['kelas'], info['sheets']
            gnum = extract_kelas_num(kelas, kelas_fmt)
            if gnum not in gugus_set: continue  # class not in this REKAP -> out of scope
            # Track which roster name produced the best score too, not just the
            # score itself — a low score alone is a dead end; the closest
            # candidate's exact (gugus, row) lets an admin confirm specific
            # subtests for this match straight from the review UI (a match
            # calc_match couldn't clear on its own, e.g. a vowel-dropped
            # abbreviation), instead of only offering a hint with no action.
            best, best_candidate, cand_gugus, cand_row = 0, None, None, None
            for gn2, rows in gugus_rows.items():
                for ri, rn in rows.items():
                    sc = calc_match(k, rn)
                    if sc > best: best, best_candidate, cand_gugus, cand_row = sc, rn, gn2, ri
            if best >= threshold: continue      # student is represented somewhere
            sheet_list = ', '.join(sorted(sheets))
            confirmable = best >= LEFTOVER_CANDIDATE_MIN and best_candidate is not None
            leftover.append((gnum, display, kelas, round(best, 3), sheet_list, best_candidate))
            # closest_candidate is a separate structured field (not folded into
            # `detail`) so the dashboard can render it on its own, in English,
            # without duplicating/mixing it into this Indonesian sentence.
            detail = (f'Ada di RAW (kelas {kelas}) tapi tidak ada di roster Gugus {gnum} '
                      f'— muncul di sheet: {sheet_list}')
            pending.append({
                'id': len(pending), 'gugus': gnum, 'subtes': sheet_list or '-', 'kind': 'leftover',
                'nama_rekap': '(tidak ada di rekap)', 'nama_raw': str(display),
                'score': round(best, 3), 'source': 'raw', 'closest_candidate': best_candidate,
                'reason': 'N7_leftover', 'severity': 'warn', 'detail': detail,
                'confirmable': confirmable,
                # Internal-only (stripped from /status by the leading underscore
                # convention) — needed by submit_review to write a per-subtest
                # confirmation into the candidate's own roster row.
                '_norm_key': k, '_sheets': sorted(sheets),
                '_candidate_gugus': cand_gugus if confirmable else None,
                '_candidate_row': cand_row if confirmable else None,
            })
        log['leftover_raw'] = [{'gugus': g, 'nama': d, 'kelas': str(kl), 'best': b, 'subtes': sl,
                                 'closest_candidate': cn}
                               for g, d, kl, b, sl, cn in leftover]

        ctx = {'wb':wb,'headers':headers,'name_col':name_col,'gugus_rows':gugus_rows,
               'gugus_nums':gugus_nums,'rekap_subtes':rekap_subtes,'raw_path':raw_path,
               'subtes_sheets':subtes_sheets,'all_raw':all_raw,'all_proktor':all_proktor,
               'kelas_fmt':kelas_fmt,'tgl_pemeriksaan':tgl_pemeriksaan,'pendidikan':pendidikan,
               'log':log,'threshold':threshold,'norm_ov':norm_ov,'ge_qcols':ge_qcols,
               'header_row':header_row,'data_start':data_start,'answer_sections':answer_sections}
        if pending:
            job['_ctx'] = ctx; job['pending'] = pending
            job['status'] = 'awaiting_review'; job['progress'] = 78
            return
        job['_ctx'] = ctx
        finalize_job(job_id)
    except Exception as e:
        job['status'] = 'error'; job['error'] = f'{type(e).__name__}: {str(e)}'
        _free_job_memory(job)


def finalize_job(job_id):
    job = jobs[job_id]
    ctx = job['_ctx']
    wb = ctx['wb']; headers = ctx['headers']; name_col = ctx['name_col']
    gugus_rows = ctx['gugus_rows']; gugus_nums = ctx['gugus_nums']; rekap_subtes = ctx['rekap_subtes']
    raw_path = ctx['raw_path']; subtes_sheets = ctx['subtes_sheets']
    all_raw = ctx['all_raw']; all_proktor = ctx['all_proktor']
    kelas_fmt = ctx['kelas_fmt']; tgl_pemeriksaan = ctx['tgl_pemeriksaan']; pendidikan = ctx['pendidikan']
    log = ctx['log']; threshold = ctx['threshold']; norm_ov = ctx['norm_ov']
    data_start = ctx.get('data_start', 2)
    answer_sections = ctx.get('answer_sections') or {}
    try:
        # ── Identity ──
        job['status'] = 'Mengisi identitas...'; job['progress'] = 80
        id_sheet = find_id_sheet(raw_path, subtes_sheets)
        if id_sheet and id_sheet in all_raw:
            # Combine normal + proktor entries for identity search
            rid = pd.concat([all_raw[id_sheet], all_proktor.get(id_sheet, pd.DataFrame())], ignore_index=True)
            rid = rid.drop_duplicates(subset=['_norm'], keep='first')
            id_map = {}
            for c in rid.columns:
                u = str(c).upper().strip()
                if 'JENIS KELAMIN' in u and 'JK' in headers: id_map[c]='JK'
                elif 'TEMPAT LAHIR' in u and 'TEMPAT LAHIR' in headers: id_map[c]='TEMPAT LAHIR'
                elif 'TANGGAL LAHIR' in u and 'TGL LAHIR' in headers: id_map[c]='TGL LAHIR'
            for gnum in gugus_nums:
                if gnum not in gugus_rows: continue
                ws = wb[f'Gugus {gnum}']
                for row_idx, gn in gugus_rows[gnum].items():
                    mt = norm_ov.get(gn); best_s,best_i=0,None
                    for ri, r in rid.iterrows():
                        s = calc_match(gn, r['_norm'])
                        if mt and calc_match(mt, r['_norm'])>0.9: s=0.95
                        if s>best_s: best_s,best_i=s,ri
                    if best_s<threshold or best_i is None: continue
                    r = rid.loc[best_i]
                    for rc, rkc in id_map.items():
                        if rc in r.index and rkc in headers:
                            v = r[rc]
                            if pd.notna(v):
                                cur = ws.cell(row=row_idx, column=headers[rkc]).value
                                if cur is None or str(cur).strip()=='':
                                    # Normalize dates
                                    if rkc == 'TGL LAHIR':
                                        v = normalize_date(v) or v
                                    ws.cell(row=row_idx, column=headers[rkc], value=v)
                                    log['identity_filled'] += 1

        # ── Fix existing dates + fill uniform fields ──
        job['status'] = 'Memperbaiki format...'; job['progress'] = 88
        tgl_lahir_col = headers.get('TGL LAHIR')
        tgl_periksa_col = headers.get('TGL PEMERIKSAAN')
        pendidikan_col = headers.get('PENDIDIKAN')

        for gnum in gugus_nums:
            ws = wb[f'Gugus {gnum}']
            for row in range(data_start, ws.max_row+1):
                nm = ws.cell(row=row, column=name_col).value
                if not nm: continue
                # Fix TGL LAHIR format
                if tgl_lahir_col:
                    v = ws.cell(row=row, column=tgl_lahir_col).value
                    if v is not None:
                        fixed = normalize_date(v)
                        if fixed: ws.cell(row=row, column=tgl_lahir_col, value=fixed)
                # Fill TGL PEMERIKSAAN (uniform)
                if tgl_periksa_col and tgl_pemeriksaan:
                    ws.cell(row=row, column=tgl_periksa_col, value=tgl_pemeriksaan)
                # Fill PENDIDIKAN (uniform)
                if pendidikan_col and pendidikan:
                    ws.cell(row=row, column=pendidikan_col, value=pendidikan)

        # ── Missing-value fill + yellow highlight + stats ──
        # A student is "matched" if any Score subtest OR any answer block is
        # filled. Matched students get their empty Score cells written as 0
        # (SE…WU + computed GE); answer blocks (EPPS/RIASEC/GB) keep true blanks.
        # Yellow now means fully unmatched (nothing anywhere), not "empty score".
        job['status'] = 'Finalisasi...'; job['progress'] = 93
        subtes_cols = [headers[s] for s in rekap_subtes if s in headers]
        answer_cols = {section: list(qmap.values())
                       for section, qmap in answer_sections.items()}
        yf = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for gnum in gugus_nums:
            ws = wb[f'Gugus {gnum}']; ts,fc,yc=0,0,0
            yellow_names = []
            ans_cov = {section: 0 for section in answer_cols}
            for row in range(data_start, ws.max_row+1):
                nm = ws.cell(row=row, column=name_col).value
                if not nm: continue
                ts += 1
                # genuine coverage, measured before any 0-fill
                rf = sum(1 for c in subtes_cols if ws.cell(row=row,column=c).value is not None)
                filled_sections = [s for s, cols in answer_cols.items()
                                   if any(ws.cell(row=row,column=c).value not in (None,'') for c in cols)]
                for s in filled_sections: ans_cov[s] += 1
                fc += rf
                matched = rf > 0 or bool(filled_sections)
                if not matched:
                    for col in range(1, ws.max_column+1): ws.cell(row=row, column=col).fill = yf
                    yc += 1; log['yellow_count'] += 1
                    yellow_names.append(str(nm))
                    continue
                # report genuinely-missing Score subtests (before they become 0)
                if rf < len(subtes_cols):
                    gn_name = normalize_name(nm)
                    missing = [rekap_subtes[j] for j,c in enumerate(subtes_cols) if ws.cell(row=row,column=c).value is None]
                    log['unmatched_summary'].append({'gugus':gnum,'nama':gn_name,'missing':missing,'all_missing':False})
                # 0-fill empty Score cells (matched students only)
                for c in subtes_cols:
                    if ws.cell(row=row,column=c).value is None:
                        ws.cell(row=row, column=c, value=0)
            tc = ts * len(subtes_cols)
            log['per_gugus'][f'Gugus {gnum}'] = {'siswa':ts,'filled':fc,'total':tc,
                'pct':fc*100//tc if tc else 0,'yellow':yc,
                'answer_coverage':ans_cov}
            if yellow_names:
                log['yellow_detail'].append({'gugus':gnum, 'names':yellow_names})

        # job_id in the name so two jobs finishing in the same second don't collide.
        out_name = f"Rekap_Otomatis_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{job_id}.xlsx"
        out_path = os.path.join(UPLOAD_FOLDER, out_name)
        wb.save(out_path)
        job['progress'] = 100; job['status'] = 'done'; job['log'] = log; job['download_filename'] = out_name
        _free_job_memory(job)
    except Exception as e:
        job['status'] = 'error'; job['error'] = f'{type(e).__name__}: {str(e)}'
        _free_job_memory(job)


# ── Routes ──
# No HTML UI is served here anymore — the Ordinat Dashboard is the only
# frontend. Every route below requires the shared service token.

@app.route('/healthz')
def healthz():
    """Unauthenticated on purpose — platform health checks (Render etc.) hit
    this before any token exists. Reveals nothing beyond "process is alive"."""
    return jsonify({'status': 'ok'}), 200

@app.route('/console')
def dev_console():
    """Serve the local test console (test_console.html) same-origin so its API
    calls avoid CORS. Off by default to preserve the "no public UI" posture —
    opt in with ENABLE_TEST_CONSOLE=1 for local dev only. The page itself still
    carries the bearer token on every API call; this route only serves static
    HTML and exposes nothing."""
    if os.environ.get('ENABLE_TEST_CONSOLE', '').lower() not in ('1', 'true', 'yes'):
        return "Not found", 404
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'test_console.html')
    if not os.path.exists(path):
        return "console not found", 404
    return send_file(path)

@app.route('/process', methods=['POST'])
@require_service_token
def start_process():
    if 'raw_file' not in request.files or 'rekap_file' not in request.files:
        return jsonify({'error':'Upload kedua file'}), 400
    rf, rkf = request.files['raw_file'], request.files['rekap_file']
    # Prefix stored files with the job id so concurrent uploads that happen to
    # share a filename (e.g. two admins both uploading "RAW.xlsx") never
    # overwrite each other on disk while a background thread is still reading.
    jid = str(uuid.uuid4())[:8]
    rp = os.path.join(UPLOAD_FOLDER, f"{jid}_raw_{secure_filename(rf.filename)}")
    rkp = os.path.join(UPLOAD_FOLDER, f"{jid}_rekap_{secure_filename(rkf.filename)}")
    rf.save(rp); rkf.save(rkp)
    ov = {}
    try: ov = json.loads(request.form.get('manual_overrides','{}'))
    except: pass
    th = float(request.form.get('threshold', 0.78))
    tgl_p = request.form.get('tgl_pemeriksaan', '').strip() or None
    pend = request.form.get('pendidikan', '').strip() or None
    jobs[jid] = {'status':'queued','progress':0,'log':None,'download_filename':None,
                 'error':None,'_raw_path':rp,'_rekap_path':rkp}
    _evict_old_jobs()
    threading.Thread(target=process_job, args=(jid, rp, rkp, ov, th, tgl_p, pend), daemon=True).start()
    return jsonify({'job_id': jid})

@app.route('/status/<job_id>')
@require_service_token
def job_status(job_id):
    j = jobs.get(job_id)
    if not j: return jsonify({'error':'Job not found'}), 404
    resp = {'status':j['status'],'progress':j['progress'],'log':j.get('log'),
            'download_filename':j.get('download_filename'),'error':j.get('error')}
    if j.get('status') == 'awaiting_review':
        resp['pending'] = [{k:v for k,v in p.items() if not k.startswith('_')} for p in j.get('pending', [])]
    return jsonify(resp)

@app.route('/review/<job_id>', methods=['POST'])
@require_service_token
def submit_review(job_id):
    j = jobs.get(job_id)
    if not j: return jsonify({'error':'Job not found'}), 404
    if j.get('status') != 'awaiting_review':
        return jsonify({'error':'Job tidak sedang menunggu review'}), 400
    data = request.get_json(silent=True) or {}
    decisions = data.get('decisions', {})
    ctx = j['_ctx']; wb = ctx['wb']; log = ctx['log']
    headers = ctx.get('headers', {}); answer_sections = ctx.get('answer_sections') or {}
    ge_qcols = ctx.get('ge_qcols', []); raw_path = ctx.get('raw_path')
    for item in j.get('pending', []):
        dec = decisions.get(str(item['id']))
        if not dec: continue
        kind = item.get('kind')
        if kind == 'leftover':
            # Unlike value/answer_choice (boolean per item), a leftover item's
            # decision is the LIST of subtest codes the admin checked in the
            # review UI's per-subtest dropdown — not confirmed wholesale, since
            # a candidate name plausible enough to suggest isn't automatically
            # plausible enough to blindly apply to every one of a student's
            # subtests. Only 'confirmable' items (see process_job) have a
            # resolved candidate row to write into at all.
            if not item.get('confirmable') or not isinstance(dec, list) or not dec:
                continue
            if _apply_leftover_confirm(wb, raw_path, item, dec, headers, answer_sections, ge_qcols):
                log['total_scores'] += 1
                log['cross_kelas'] += 1  # always lands on a different roster row than the one RAW claimed
            continue
        ws = wb[f"Gugus {item['gugus']}"]
        if kind == 'answer_choice':
            for col, val in zip(item['_targets'], item['_answers']):
                if col and val is not None:
                    ws.cell(row=item['_row_idx'], column=col, value=val)
        else:
            ws.cell(row=item['_row_idx'], column=item['_col_idx'], value=item['_value'])
        log['total_scores'] += 1
        if item['source'] != 'own': log['cross_kelas'] += 1
    j['pending'] = []
    j['status'] = 'Melanjutkan proses...'; j['progress'] = 80
    threading.Thread(target=finalize_job, args=(job_id,), daemon=True).start()
    return jsonify({'ok': True})

@app.route('/download/<filename>')
@require_service_token
def download(filename):
    p = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
    if not os.path.exists(p): return "File not found", 404
    return send_file(p, as_attachment=True, download_name=filename)

if __name__ == '__main__':
    # debug=True exposes Werkzeug's interactive debugger (an RCE surface on any
    # reachable instance) — never on by default. Opt in explicitly for local dev.
    debug = os.environ.get('FLASK_DEBUG', '').lower() in ('1', 'true', 'yes')
    app.run(debug=debug, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
