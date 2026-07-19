"""The matching / review / finalize pipeline.

`process_job` reads the RAW workbook, matches every roster student across three
tiers (own class -> cross class -> proktor catch-all) for both score and
answer-choice subtests, recomputes GE, and either writes confident matches
straight into the REKAP workbook or queues anomalies for manual review.
`finalize_job` then backfills identity fields, normalizes dates, 0-fills matched
students, highlights the unmatched, and saves the output workbook.

Behavior is identical to the original single-file version. The performance work
here is confined to three result-preserving changes:
  1. `calc_match` is memoized (see matching), so the same name pair is scored
     once instead of once per subtest/pass.
  2. RAW candidate pools are pre-extracted to (index, name) tuples once per
     group, replacing per-student DataFrame.iterrows().
  3. The leftover roster scan early-exits as soon as a student is shown to be
     represented, instead of always computing a global maximum.
"""
from __future__ import annotations

import os
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .config import AUTO_CONFIDENCE, LEFTOVER_CANDIDATE_MIN, RAW_TO_SECTION
from .dates import normalize_date
from .jobstore import UPLOAD_FOLDER, free_job_memory, jobs
from .layout import detect_header_layout, map_answer_sections, read_field_headers
from .matching import (
    as_pairs,
    best_name_match,
    calc_match,
    closest_roster_candidate,
    normalize_name,
)
from .quality import assess_answer_quality, assess_ge
from .raw_reader import (
    detect_answer_sheets,
    detect_kelas_fmt,
    extract_kelas_num,
    fast_detect_sheets,
    find_id_sheet,
    read_ge,
    read_raw_answers,
    read_raw_minimal,
)
from .scoring import extract_choice, value_for


def queue_or_write(ws, row_idx, col_idx, sv, best_s, gnum, subtes, gn_display,
                   matched_name, source, log, pending, flag=None) -> bool:
    """Write score directly if high-confidence and clean; otherwise queue it for
    manual review. `flag` is an optional (reason, severity, detail) anomaly (e.g.
    GE zero-signal) that forces review even on a confident name match.
    Returns True if the row is considered handled (found), either way."""
    if pd.isna(sv): return False
    # Scores come from Google Forms responses — a stray non-numeric cell must
    # skip just this row (it'll surface as missing/yellow), never crash the job.
    try:
        score_val = int(float(sv))
    except (ValueError, TypeError):
        return False
    if flag is None and best_s >= AUTO_CONFIDENCE:
        ws.cell(row=row_idx, column=col_idx, value=score_val)
        log['total_scores'] += 1
        if source != 'own': log['cross_kelas'] += 1
    else:
        reason, severity, detail = flag if flag else (
            'N1_borderline', 'block', f'skor kecocokan nama {round(best_s, 3)}')
        pending.append({
            'id': len(pending), 'gugus': gnum, 'subtes': subtes,
            'nama_rekap': gn_display, 'nama_raw': str(matched_name),
            'score': round(best_s, 3), 'source': source, 'kind': 'value',
            'reason': reason, 'severity': severity, 'detail': detail,
            '_row_idx': row_idx, '_col_idx': col_idx, '_value': score_val,
        })
    return True


def _apply_leftover_confirm(wb, raw_path, item, subtes_list, headers,
                            answer_sections, ge_qcols) -> bool:
    """Write a confirmed subset of a leftover student's subtests into their
    matched candidate's roster row (item['_candidate_gugus']/['_candidate_row'],
    set only when 'confirmable' — see the leftover-check pass in process_job).

    Re-reads raw_path directly for just the requested sheets rather than caching
    every RAW row up front — leftover confirmation is the rare, manually-
    triggered path, not the per-cell hot loop the rest of this file optimizes
    for. raw_path is still on disk at this point: it's only removed by
    free_job_memory, which runs after finalize_job, which this function's caller
    (submit_review) always runs strictly before.

    Returns True if at least one subtest was actually written."""
    gnum = item.get('_candidate_gugus')
    row_idx = item.get('_candidate_row')
    k = item.get('_norm_key')
    if gnum is None or row_idx is None or not k: return False
    ws = wb[f'Gugus {gnum}']
    wrote = False
    for sn in subtes_list:
        sn = str(sn).strip().upper()
        if sn in headers:
            # Score-type subtest (SE/WA/AN/GE/RA/ZR/FA/WU/ME/...): one cell.
            if sn == 'GE':
                df, gqc = read_ge(raw_path, sn)
            else:
                df, gqc = read_raw_minimal(raw_path, sn), []
            df = df.copy()
            df['_norm'] = df['NAMA LENGKAP'].apply(normalize_name)
            match = df[df['_norm'] == k]
            if match.empty: continue
            row = match.iloc[0]
            sv = value_for(sn, row, gqc)
            if pd.isna(sv): continue
            try:
                score_val = int(float(sv))
            except (ValueError, TypeError):
                continue
            ws.cell(row=row_idx, column=headers[sn], value=score_val)
            wrote = True
        else:
            # Answer-choice subtest (EPPS/RIASEC/GB): a whole question block.
            section = RAW_TO_SECTION.get(sn, sn)
            qmap = answer_sections.get(section)
            if not qmap: continue
            found_sheet = next((asn for asn, _ in detect_answer_sheets(raw_path)
                                if asn.strip().upper() == sn), None)
            if not found_sheet: continue
            adf, qcols = read_raw_answers(raw_path, found_sheet)
            match = adf[adf['_norm'] == k]
            if match.empty: continue
            row = match.iloc[0]
            n_write = min(len(qcols), len(qmap))
            answers = [extract_choice(row[qcols[i]]) for i in range(n_write)]
            targets = [qmap.get(i + 1) for i in range(n_write)]
            for col, val in zip(targets, answers):
                if col and val is not None:
                    ws.cell(row=row_idx, column=col, value=val)
                    wrote = True
    return wrote


def process_job(job_id: str, raw_path: str, rekap_path: str, overrides: dict[str, str],
                threshold: float, tgl_pemeriksaan: str | None, pendidikan: str | None) -> None:
    """Background worker: read RAW, match every roster student, write confident
    matches, and queue anomalies. Finalizes inline when nothing needs review."""
    # Start each job from a clean match cache so a completed job's entries don't
    # linger in memory; within this job the cache still absorbs every repeat.
    calc_match.cache_clear()
    job = jobs[job_id]
    try:
        job['status'] = 'Mendeteksi sheet...'; job['progress'] = 3
        norm_ov = {normalize_name(k): (normalize_name(v) if v else None)
                   for k, v in overrides.items()}
        subtes_sheets, _ = fast_detect_sheets(raw_path)
        if not subtes_sheets:
            job['status'] = 'error'; job['error'] = 'Tidak ada sheet subtes di RAW'; return

        job['progress'] = 8
        all_raw: dict[str, pd.DataFrame] = {}
        all_proktor: dict[str, pd.DataFrame] = {}
        kelas_fmt = None
        ge_qcols: list[object] = []
        raw_roster: dict[str, dict[str, Any]] = {}  # _norm -> {'nama','kelas','sheets'}:
        #                  every real RAW student (+ which sheets), for the leftover check
        for i, sn in enumerate(subtes_sheets):
            job['status'] = f'Membaca {sn}...'
            if sn.strip().upper() == 'GE':
                raw, ge_qcols = read_ge(raw_path, sn)  # keep answer cols to recompute
            else:
                raw = read_raw_minimal(raw_path, sn)
            raw['_norm'] = raw['NAMA LENGKAP'].apply(normalize_name)
            # Split: normal kelas vs Proktor entries
            is_proktor = raw['KELAS'].astype(str).str.contains('Proktor', na=False)
            proktor_df = raw[is_proktor].drop_duplicates(subset=['_norm'], keep='first')
            normal_df = raw[~is_proktor].drop_duplicates(subset=['_norm', 'KELAS'], keep='first')
            all_raw[sn] = normal_df
            all_proktor[sn] = proktor_df
            for k, nm, kl in zip(normal_df['_norm'], normal_df['NAMA LENGKAP'], normal_df['KELAS']):
                if not k: continue
                raw_roster.setdefault(k, {'nama': nm, 'kelas': kl, 'sheets': set()})['sheets'].add(sn)
            if kelas_fmt is None:
                kv = [v for v in normal_df['KELAS'].dropna().unique()]
                kelas_fmt = detect_kelas_fmt(kv)
            job['progress'] = 8 + int(22 * (i + 1) / len(subtes_sheets))

        job['status'] = 'Membaca Rekap...'; job['progress'] = 32
        wb = load_workbook(rekap_path)
        gsheets = sorted([s for s in wb.sheetnames if s.lower().startswith('gugus')],
                         key=lambda x: int(''.join(filter(str.isdigit, x)) or 0))
        if not gsheets:
            job['status'] = 'error'; job['error'] = 'Tidak ada sheet Gugus'; return

        ws0 = wb[gsheets[0]]
        header_row, data_start = detect_header_layout(ws0)
        headers = read_field_headers(ws0, header_row)
        answer_sections = map_answer_sections(ws0, header_row)
        name_col = headers.get('NAMA', 2)
        rekap_subtes = [s for s in subtes_sheets if s in headers]

        gugus_rows: dict[int, dict[int, str]] = {}
        gugus_all_names: dict[int, list[str]] = {}
        gugus_nums: list[int] = []
        for sn in gsheets:
            try:
                gnum = int(''.join(filter(str.isdigit, sn)))
            except Exception:
                continue
            gugus_nums.append(gnum)
            ws = wb[sn]
            gugus_rows[gnum] = {}
            for row in range(data_start, ws.max_row + 1):
                cv = ws.cell(row=row, column=name_col).value
                if cv: gugus_rows[gnum][row] = normalize_name(cv)
            gugus_all_names[gnum] = list(gugus_rows[gnum].values())

        job['progress'] = 38; job['status'] = 'Mencocokkan nama...'
        log: dict[str, Any] = {
            'total_scores': 0, 'cross_kelas': 0, 'identity_filled': 0, 'yellow_count': 0,
            'per_gugus': {}, 'unmatched_summary': [], 'yellow_detail': [],
            'subtes_detected': subtes_sheets, 'rekap_subtes': rekap_subtes}

        total_steps = len(rekap_subtes) * len(gugus_nums); step = 0
        pending: list[dict[str, Any]] = []
        for subtes in rekap_subtes:
            raw = all_raw[subtes]; proktor = all_proktor[subtes]; col_idx = headers[subtes]
            is_ge = str(subtes).strip().upper() == 'GE'
            # Proktor pool is the same for every gugus of this subtest — extract once.
            proktor_pairs = as_pairs(proktor)
            for gnum in gugus_nums:
                if gnum not in gugus_rows: step += 1; continue
                kstr = f'{kelas_fmt} {gnum}' if kelas_fmt != 'XI.' else f'XI.{gnum}'
                ws = wb[f'Gugus {gnum}']
                # Own- and cross-class pools depend only on the gugus, not the
                # student — extract them once here instead of per roster row.
                raw_own = raw[raw['KELAS'] == kstr]
                raw_other = raw[raw['KELAS'] != kstr]
                own_pairs = as_pairs(raw_own)
                other_pairs = as_pairs(raw_other)
                for row_idx, gn in gugus_rows[gnum].items():
                    found = False; mt = norm_ov.get(gn)
                    # Step 1: own kelas
                    best_s, best_i = best_name_match(gn, mt, own_pairs)
                    if best_s >= threshold and best_i is not None:
                        mrow = raw.loc[best_i]
                        sv = value_for(subtes, mrow, ge_qcols)
                        flag = assess_ge(sv, mrow, ge_qcols) if is_ge else None
                        found = queue_or_write(ws, row_idx, col_idx, sv, best_s, gnum,
                                               subtes, gn, mrow['NAMA LENGKAP'], 'own', log, pending, flag)
                    # Step 2: cross-kelas
                    if not found:
                        best_s, best_i = best_name_match(gn, mt, other_pairs)
                        if best_s >= threshold and best_i is not None:
                            rr = raw.loc[best_i]; rg = extract_kelas_num(rr['KELAS'], kelas_fmt)
                            claimed = rg in gugus_all_names and any(
                                calc_match(og, rr['_norm']) >= threshold for og in gugus_all_names[rg])
                            if not claimed:
                                sv = value_for(subtes, rr, ge_qcols)
                                flag = assess_ge(sv, rr, ge_qcols) if is_ge else None
                                found = queue_or_write(ws, row_idx, col_idx, sv, best_s, gnum,
                                                       subtes, gn, rr['NAMA LENGKAP'], 'cross', log, pending, flag)
                    # Step 3: search Proktor entries (students who picked wrong kelas option)
                    if not found and len(proktor) > 0:
                        best_s, best_i = best_name_match(gn, mt, proktor_pairs)
                        if best_s >= threshold and best_i is not None:
                            mrow = proktor.loc[best_i]
                            sv = value_for(subtes, mrow, ge_qcols)
                            flag = assess_ge(sv, mrow, ge_qcols) if is_ge else None
                            found = queue_or_write(ws, row_idx, col_idx, sv, best_s, gnum,
                                                   subtes, gn, mrow['NAMA LENGKAP'], 'proktor', log, pending, flag)
                step += 1
                job['progress'] = min(38 + int(40 * step / max(total_steps, 1)), 78)
            job['status'] = f'Mencocokkan {subtes}...'

        # ── Answer-choice subtests (EPPS / RIASEC / GB) ──
        # Same 3-tier name match as above, but write parsed choice tokens across
        # the section's question-number columns instead of one Score cell.
        if answer_sections:
            job['status'] = 'Mencocokkan jawaban...'
            for sn, section in detect_answer_sheets(raw_path):
                if section not in answer_sections: continue
                adf, qcols = read_raw_answers(raw_path, sn)
                is_prok = adf['KELAS'].astype(str).str.contains('Proktor', na=False)
                aprok = adf[is_prok].drop_duplicates(subset=['_norm'], keep='first')
                anorm = adf[~is_prok].drop_duplicates(subset=['_norm', 'KELAS'], keep='first')
                for k, nm, kl in zip(anorm['_norm'], anorm['NAMA LENGKAP'], anorm['KELAS']):
                    if not k: continue
                    raw_roster.setdefault(k, {'nama': nm, 'kelas': kl, 'sheets': set()})['sheets'].add(sn)
                qmap = answer_sections[section]
                n_write = min(len(qcols), len(qmap))
                aprok_pairs = as_pairs(aprok)  # same across every gugus
                for gnum in gugus_nums:
                    if gnum not in gugus_rows: continue
                    kstr = f'{kelas_fmt} {gnum}' if kelas_fmt != 'XI.' else f'XI.{gnum}'
                    ws = wb[f'Gugus {gnum}']
                    own = anorm[anorm['KELAS'] == kstr]; other = anorm[anorm['KELAS'] != kstr]
                    own_pairs = as_pairs(own); other_pairs = as_pairs(other)
                    for row_idx, gn in gugus_rows[gnum].items():
                        mt = norm_ov.get(gn)
                        best_s, best_i, src, pool = 0, None, None, None
                        bs, bi = best_name_match(gn, mt, own_pairs)
                        if bs >= threshold and bi is not None:
                            best_s, best_i, src, pool = bs, bi, 'own', anorm
                        if best_i is None:
                            bs, bi = best_name_match(gn, mt, other_pairs)
                            if bs >= threshold and bi is not None:
                                rr = anorm.loc[bi]; rg = extract_kelas_num(rr['KELAS'], kelas_fmt)
                                claimed = rg in gugus_all_names and any(
                                    calc_match(og, rr['_norm']) >= threshold for og in gugus_all_names[rg])
                                if not claimed: best_s, best_i, src, pool = bs, bi, 'cross', anorm
                        if best_i is None and len(aprok) > 0:
                            bs, bi = best_name_match(gn, mt, aprok_pairs)
                            if bs >= threshold and bi is not None:
                                best_s, best_i, src, pool = bs, bi, 'proktor', aprok
                        if best_i is None: continue
                        row = pool.loc[best_i]
                        answers = [extract_choice(row[qcols[i]]) for i in range(n_write)]
                        targets = [qmap.get(i + 1) for i in range(n_write)]
                        # Quality gate: even a confident name match must be
                        # confirmed if the content looks ngawur.
                        flag = assess_answer_quality(answers)
                        if flag is None and best_s >= AUTO_CONFIDENCE:
                            for col, val in zip(targets, answers):
                                if col and val is not None: ws.cell(row=row_idx, column=col, value=val)
                            log['total_scores'] += 1
                            if src != 'own': log['cross_kelas'] += 1
                        else:
                            reason, severity, detail = flag if flag else (
                                'N1_borderline', 'block', f'skor kecocokan nama {round(best_s, 3)}')
                            pending.append({
                                'id': len(pending), 'gugus': gnum, 'subtes': section,
                                'nama_rekap': gn, 'nama_raw': str(row['NAMA LENGKAP']),
                                'score': round(best_s, 3), 'source': src, 'kind': 'answer_choice',
                                'reason': reason, 'severity': severity, 'detail': detail,
                                '_row_idx': row_idx, '_answers': answers, '_targets': targets,
                            })

        # ── Leftover check (N7): RAW student missing from the REKAP roster ──
        # A RAW student counts as "missing" only if their class maps to a gugus
        # that exists in this REKAP yet no roster name matches them — otherwise
        # they belong to a class this REKAP simply doesn't cover (not an anomaly).
        gugus_set = set(gugus_nums)
        leftover = []
        for k, info in raw_roster.items():
            display, kelas, sheets = info['nama'], info['kelas'], info['sheets']
            gnum = extract_kelas_num(kelas, kelas_fmt)
            if gnum not in gugus_set: continue  # class not in this REKAP -> out of scope
            # Track which roster name produced the best score too, not just the
            # score itself — a low score alone is a dead end; the closest
            # candidate's exact (gugus, row) lets an admin confirm specific
            # subtests for this match straight from the review UI (a match
            # calc_match couldn't clear on its own, e.g. a vowel-dropped
            # abbreviation), instead of only offering a hint with no action.
            # Early-exits the moment a >= threshold roster name is found.
            represented, best, best_candidate, cand_gugus, cand_row = closest_roster_candidate(
                k, gugus_rows, threshold)
            if represented: continue      # student is represented somewhere
            sheet_list = ', '.join(sorted(sheets))
            confirmable = best >= LEFTOVER_CANDIDATE_MIN and best_candidate is not None
            leftover.append((gnum, display, kelas, round(best, 3), sheet_list, best_candidate))
            # closest_candidate is a separate structured field (not folded into
            # `detail`) so the dashboard can render it on its own, in English,
            # without duplicating/mixing it into this Indonesian sentence.
            detail = (f'Ada di RAW (kelas {kelas}) tapi tidak ada di roster Gugus {gnum} '
                      f'— muncul di sheet: {sheet_list}')
            pending.append({
                'id': len(pending), 'gugus': gnum, 'subtes': sheet_list or '-', 'kind': 'leftover',
                'nama_rekap': '(tidak ada di rekap)', 'nama_raw': str(display),
                'score': round(best, 3), 'source': 'raw', 'closest_candidate': best_candidate,
                'reason': 'N7_leftover', 'severity': 'warn', 'detail': detail,
                'confirmable': confirmable,
                # Internal-only (stripped from /status by the leading underscore
                # convention) — needed by submit_review to write a per-subtest
                # confirmation into the candidate's own roster row.
                '_norm_key': k, '_sheets': sorted(sheets),
                '_candidate_gugus': cand_gugus if confirmable else None,
                '_candidate_row': cand_row if confirmable else None,
            })
        log['leftover_raw'] = [{'gugus': g, 'nama': d, 'kelas': str(kl), 'best': b, 'subtes': sl,
                                'closest_candidate': cn}
                               for g, d, kl, b, sl, cn in leftover]

        ctx = {'wb': wb, 'headers': headers, 'name_col': name_col, 'gugus_rows': gugus_rows,
               'gugus_nums': gugus_nums, 'rekap_subtes': rekap_subtes, 'raw_path': raw_path,
               'subtes_sheets': subtes_sheets, 'all_raw': all_raw, 'all_proktor': all_proktor,
               'kelas_fmt': kelas_fmt, 'tgl_pemeriksaan': tgl_pemeriksaan, 'pendidikan': pendidikan,
               'log': log, 'threshold': threshold, 'norm_ov': norm_ov, 'ge_qcols': ge_qcols,
               'header_row': header_row, 'data_start': data_start, 'answer_sections': answer_sections}
        if pending:
            job['_ctx'] = ctx; job['pending'] = pending
            job['status'] = 'awaiting_review'; job['progress'] = 78
            return
        job['_ctx'] = ctx
        finalize_job(job_id)
    except Exception as e:
        job['status'] = 'error'; job['error'] = f'{type(e).__name__}: {str(e)}'
        free_job_memory(job)


def finalize_job(job_id: str) -> None:
    """Second phase: backfill identity fields, normalize dates + uniform fields,
    0-fill matched students, highlight the fully-unmatched, and save the output
    workbook. Runs inline from process_job or after a review is submitted."""
    job = jobs[job_id]
    ctx = job['_ctx']
    wb = ctx['wb']; headers = ctx['headers']; name_col = ctx['name_col']
    gugus_rows = ctx['gugus_rows']; gugus_nums = ctx['gugus_nums']; rekap_subtes = ctx['rekap_subtes']
    raw_path = ctx['raw_path']; subtes_sheets = ctx['subtes_sheets']
    all_raw = ctx['all_raw']; all_proktor = ctx['all_proktor']
    kelas_fmt = ctx['kelas_fmt']; tgl_pemeriksaan = ctx['tgl_pemeriksaan']; pendidikan = ctx['pendidikan']
    log = ctx['log']; threshold = ctx['threshold']; norm_ov = ctx['norm_ov']
    data_start = ctx.get('data_start', 2)
    answer_sections = ctx.get('answer_sections') or {}
    try:
        # ── Identity ──
        job['status'] = 'Mengisi identitas...'; job['progress'] = 80
        id_sheet = find_id_sheet(raw_path, subtes_sheets)
        if id_sheet and id_sheet in all_raw:
            # Combine normal + proktor entries for identity search
            rid = pd.concat([all_raw[id_sheet], all_proktor.get(id_sheet, pd.DataFrame())],
                            ignore_index=True)
            rid = rid.drop_duplicates(subset=['_norm'], keep='first')
            id_map = {}
            for c in rid.columns:
                u = str(c).upper().strip()
                if 'JENIS KELAMIN' in u and 'JK' in headers: id_map[c] = 'JK'
                elif 'TEMPAT LAHIR' in u and 'TEMPAT LAHIR' in headers: id_map[c] = 'TEMPAT LAHIR'
                elif 'TANGGAL LAHIR' in u and 'TGL LAHIR' in headers: id_map[c] = 'TGL LAHIR'
            # rid is fixed for the whole identity pass — extract its pool once.
            rid_pairs = as_pairs(rid)
            for gnum in gugus_nums:
                if gnum not in gugus_rows: continue
                ws = wb[f'Gugus {gnum}']
                for row_idx, gn in gugus_rows[gnum].items():
                    mt = norm_ov.get(gn)
                    best_s, best_i = best_name_match(gn, mt, rid_pairs)
                    if best_s < threshold or best_i is None: continue
                    r = rid.loc[best_i]
                    for rc, rkc in id_map.items():
                        if rc in r.index and rkc in headers:
                            v = r[rc]
                            if pd.notna(v):
                                cur = ws.cell(row=row_idx, column=headers[rkc]).value
                                if cur is None or str(cur).strip() == '':
                                    # Normalize dates
                                    if rkc == 'TGL LAHIR':
                                        v = normalize_date(v) or v
                                    ws.cell(row=row_idx, column=headers[rkc], value=v)
                                    log['identity_filled'] += 1

        # ── Fix existing dates + fill uniform fields ──
        job['status'] = 'Memperbaiki format...'; job['progress'] = 88
        tgl_lahir_col = headers.get('TGL LAHIR')
        tgl_periksa_col = headers.get('TGL PEMERIKSAAN')
        pendidikan_col = headers.get('PENDIDIKAN')

        for gnum in gugus_nums:
            ws = wb[f'Gugus {gnum}']
            for row in range(data_start, ws.max_row + 1):
                nm = ws.cell(row=row, column=name_col).value
                if not nm: continue
                # Fix TGL LAHIR format
                if tgl_lahir_col:
                    v = ws.cell(row=row, column=tgl_lahir_col).value
                    if v is not None:
                        fixed = normalize_date(v)
                        if fixed: ws.cell(row=row, column=tgl_lahir_col, value=fixed)
                # Fill TGL PEMERIKSAAN (uniform)
                if tgl_periksa_col and tgl_pemeriksaan:
                    ws.cell(row=row, column=tgl_periksa_col, value=tgl_pemeriksaan)
                # Fill PENDIDIKAN (uniform)
                if pendidikan_col and pendidikan:
                    ws.cell(row=row, column=pendidikan_col, value=pendidikan)

        # ── Missing-value fill + yellow highlight + stats ──
        # A student is "matched" if any Score subtest OR any answer block is
        # filled. Matched students get their empty Score cells written as 0
        # (SE…WU + computed GE); answer blocks (EPPS/RIASEC/GB) keep true blanks.
        # Yellow now means fully unmatched (nothing anywhere), not "empty score".
        job['status'] = 'Finalisasi...'; job['progress'] = 93
        subtes_cols = [headers[s] for s in rekap_subtes if s in headers]
        answer_cols = {section: list(qmap.values())
                       for section, qmap in answer_sections.items()}
        yf = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for gnum in gugus_nums:
            ws = wb[f'Gugus {gnum}']; ts, fc, yc = 0, 0, 0
            yellow_names = []
            ans_cov = {section: 0 for section in answer_cols}
            for row in range(data_start, ws.max_row + 1):
                nm = ws.cell(row=row, column=name_col).value
                if not nm: continue
                ts += 1
                # genuine coverage, measured before any 0-fill
                rf = sum(1 for c in subtes_cols if ws.cell(row=row, column=c).value is not None)
                filled_sections = [s for s, cols in answer_cols.items()
                                   if any(ws.cell(row=row, column=c).value not in (None, '') for c in cols)]
                for s in filled_sections: ans_cov[s] += 1
                fc += rf
                matched = rf > 0 or bool(filled_sections)
                if not matched:
                    for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).fill = yf
                    yc += 1; log['yellow_count'] += 1
                    yellow_names.append(str(nm))
                    continue
                # report genuinely-missing Score subtests (before they become 0)
                if rf < len(subtes_cols):
                    gn_name = normalize_name(nm)
                    missing = [rekap_subtes[j] for j, c in enumerate(subtes_cols)
                               if ws.cell(row=row, column=c).value is None]
                    log['unmatched_summary'].append(
                        {'gugus': gnum, 'nama': gn_name, 'missing': missing, 'all_missing': False})
                # 0-fill empty Score cells (matched students only)
                for c in subtes_cols:
                    if ws.cell(row=row, column=c).value is None:
                        ws.cell(row=row, column=c, value=0)
            tc = ts * len(subtes_cols)
            log['per_gugus'][f'Gugus {gnum}'] = {'siswa': ts, 'filled': fc, 'total': tc,
                                                 'pct': fc * 100 // tc if tc else 0, 'yellow': yc,
                                                 'answer_coverage': ans_cov}
            if yellow_names:
                log['yellow_detail'].append({'gugus': gnum, 'names': yellow_names})

        # job_id in the name so two jobs finishing in the same second don't collide.
        out_name = f"Rekap_Otomatis_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{job_id}.xlsx"
        out_path = os.path.join(UPLOAD_FOLDER, out_name)
        wb.save(out_path)
        job['progress'] = 100; job['status'] = 'done'; job['log'] = log
        job['download_filename'] = out_name
        free_job_memory(job)
    except Exception as e:
        job['status'] = 'error'; job['error'] = f'{type(e).__name__}: {str(e)}'
        free_job_memory(job)
