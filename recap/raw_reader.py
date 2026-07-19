"""RAW workbook inspection and column readers.

The RAW file is a Google-Forms export: one sheet per subtest. Score-type sheets
carry a single `Score` column; answer-type sheets (EPPS/RIASEC/GB) carry one
answer string per numbered question. These helpers detect the shape of each
sheet and load only the columns the pipeline needs.
"""
from __future__ import annotations

import re

import pandas as pd
from openpyxl import load_workbook

from .config import RAW_TO_SECTION
from .matching import normalize_name


def fast_detect_sheets(raw_path: str) -> tuple[list[str], list[str]]:
    """Split RAW sheets into (score_subtests, other) by header shape, reading
    only row 1 of each sheet."""
    wb = load_workbook(raw_path, read_only=True, data_only=True)
    subtes: list[str] = []
    other: list[str] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        hdrs = [str(h).strip() if h else "" for h in row1]
        if 'Score' in hdrs and 'NAMA LENGKAP' in hdrs and 'KELAS' in hdrs:
            subtes.append(sn)
        else:
            other.append(sn)
    wb.close()
    return subtes, other


def read_raw_minimal(raw_path: str, sheet_name: str) -> pd.DataFrame:
    """Read a score-type sheet, keeping only identity + Score columns."""
    needed = ['Timestamp', 'Score', 'NAMA LENGKAP', 'KELAS',
              'JENIS KELAMIN', 'TEMPAT LAHIR', 'TANGGAL LAHIR', 'TANGGAL TES', 'PENDIDIKAN ']
    return pd.read_excel(raw_path, sheet_name=sheet_name, usecols=lambda c: c in needed)


def detect_kelas_fmt(vals: list[object]) -> str:
    """Infer how classes are written ('KELAS X' / 'XI.' / bare 'X')."""
    for v in vals:
        s = str(v)
        if s.startswith('KELAS X'): return 'KELAS X'
        if 'XI.' in s: return 'XI.'
    return 'X'


def extract_kelas_num(kstr: object, fmt: str) -> int:
    """Pull the numeric class index out of a class label given its format."""
    try:
        s = str(kstr).strip()
        if fmt == 'KELAS X': return int(s.replace('KELAS X', '').strip())
        if fmt == 'XI.': return int(s.replace('XI.', '').strip())
        return int(s.replace('X', '').strip())
    except Exception:
        return 0


def find_id_sheet(raw_path: str, subtes_sheets: list[str]) -> str | None:
    """Pick the score sheet that carries biodata (JENIS KELAMIN), falling back
    to 'SE' then the first available sheet."""
    wb = load_workbook(raw_path, read_only=True, data_only=True)
    res = None
    for sn in subtes_sheets:
        ws = wb[sn]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        hdrs = [str(h).upper().strip() if h else "" for h in row1]
        if any('JENIS KELAMIN' in h for h in hdrs):
            res = sn; break
    wb.close()
    return res or ('SE' if 'SE' in subtes_sheets else subtes_sheets[0] if subtes_sheets else None)


def detect_answer_sheets(raw_path: str) -> list[tuple[str, str]]:
    """RAW sheets shaped as per-question answers: NAMA LENGKAP + KELAS, no
    Score, followed by numbered question columns. Returns [(sheet, section)]."""
    wb = load_workbook(raw_path, read_only=True, data_only=True)
    out: list[tuple[str, str]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        hdrs = [str(h).strip() if h else "" for h in row1]
        up = [h.upper() for h in hdrs]
        if 'NAMA LENGKAP' not in up or 'KELAS' not in up: continue
        if 'SCORE' in up: continue  # score-type -> existing path
        if not any(re.match(r'^\d+\.', h) for h in hdrs): continue
        section = RAW_TO_SECTION.get(sn.strip().upper(), sn.strip().upper())
        out.append((sn, section))
    wb.close()
    return out


def read_raw_answers(raw_path: str, sheet_name: str) -> tuple[pd.DataFrame, list[object]]:
    """Return (df, qcols): NAMA LENGKAP + KELAS + every numbered question column
    in sheet order, plus a normalized-name column."""
    df = pd.read_excel(raw_path, sheet_name=sheet_name)
    qcols = [c for c in df.columns if re.match(r'^\d+\.', str(c).strip())]
    keep = [c for c in ['NAMA LENGKAP', 'KELAS'] + qcols if c in df.columns]
    out = df[keep].copy()
    out['_norm'] = out['NAMA LENGKAP'].apply(normalize_name)
    return out, qcols


def read_ge(raw_path: str, sheet_name: str) -> tuple[pd.DataFrame, list[object]]:
    """Return (df, qcols): Score + NAMA LENGKAP + KELAS + the numbered GE answer
    columns (dropped by read_raw_minimal), so the score can be recomputed."""
    df = pd.read_excel(raw_path, sheet_name=sheet_name)
    qcols = [c for c in df.columns if re.match(r'^\d+\.', str(c).strip())]
    keep = [c for c in ['Score', 'NAMA LENGKAP', 'KELAS'] + qcols if c in df.columns]
    return df[keep].copy(), qcols
