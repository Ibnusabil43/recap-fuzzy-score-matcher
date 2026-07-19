"""REKAP template layout detection.

The v2 REKAP template uses a two-row header: row 1 holds merged section labels
(BIODATA / EPPS / RIASEC / GAYA BELAJAR) and row 2 holds the field headers
(NOMOR, NAMA, JK, the Score-subtes labels, and per-question numbers), with data
from row 3. Older single-row templates put field headers in row 1 and data from
row 2. Everything here auto-detects which layout it's looking at so the existing
Score-copy path keeps working unchanged on both.
"""
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from .config import ANSWER_SECTION_LABELS


def detect_header_layout(ws: Worksheet) -> tuple[int, int]:
    """Return (header_row, data_start) by locating the row that holds 'NAMA'.
    Two-row template -> (2, 3); legacy single-row template -> (1, 2)."""
    for hr in (1, 2):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=hr, column=col).value
            if v is not None and str(v).strip().upper() == 'NAMA':
                return hr, hr + 1
    return 1, 2  # fallback: legacy behavior


def read_field_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    """Map field-label -> column from header_row, skipping the numeric
    per-question cells (1,2,3,...) that belong to answer-section blocks."""
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            continue  # question number, not a field label
        sv = str(v).strip()
        if not sv or sv.isdigit():
            continue
        headers[sv] = col
    return headers


def map_answer_sections(ws: Worksheet, header_row: int) -> dict[str, dict[int, int]]:
    """Return {section_label: {question_num(int): column}} for the answer-type
    blocks. Section spans come from the merged label band in row 1; question
    numbers come from header_row. Empty on legacy templates that lack these
    blocks."""
    label_cols: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None and str(v).strip().upper() in ANSWER_SECTION_LABELS:
            label_cols[col] = str(v).strip().upper()
    if not label_cols:
        return {}
    # resolve each label's column span from the merged range it anchors
    spans: dict[str, tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.min_col in label_cols:
            spans[label_cols[mr.min_col]] = (mr.min_col, mr.max_col)
    # non-merged single-cell labels: span to the column before the next label/end
    starts = sorted(label_cols)
    for i, c0 in enumerate(starts):
        lbl = label_cols[c0]
        if lbl in spans:
            continue
        c1 = (starts[i + 1] - 1) if i + 1 < len(starts) else ws.max_column
        spans[lbl] = (c0, c1)
    sections: dict[str, dict[int, int]] = {}
    for lbl, (c0, c1) in spans.items():
        qmap: dict[int, int] = {}
        for col in range(c0, c1 + 1):
            hv = ws.cell(row=header_row, column=col).value
            if isinstance(hv, (int, float)) and not isinstance(hv, bool):
                qmap[int(hv)] = col
            elif hv is not None and str(hv).strip().isdigit():
                qmap[int(str(hv).strip())] = col
        if qmap:
            sections[lbl] = qmap
    return sections
